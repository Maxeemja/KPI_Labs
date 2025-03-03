from openpyxl import Workbook

from lab2 import ExpressionOptimizer, TreeBuilder
from lab3 import DistributiveTransformator
from lab4 import AssociativeTransformer
from lab5 import VLIWSystem


def simulate_parallel_calculation_in_vliw_system(expression_tree_root):
    parameters = {
        "processors_count": 6,
        "memory_bank_count": 2,
        "add_duration": 1,
        "sub_duration": 1,
        "mult_duration": 2,
        "div_duration": 4
    }

    VLIW_system = VLIWSystem(**parameters)
    VLIW_system.parallel_calculation_simulation(expression_tree_root)

    return VLIW_system.get_system_characteristics()


def writing_stats_in_xlsx(ws, expression_form, sequential_calculation_time, parallel_calculation_time,
                          acceleration_factor, system_efficiency):
    print(f"Форма виразу: {expression_form}")
    print(f"Час послідовного обрахунку виразу: {sequential_calculation_time}")
    print(f"Час паралельного обрахунку виразу: {parallel_calculation_time}")
    print(f"Коефіцієнт прискорення: {acceleration_factor}")
    print(f"Ефективність: {system_efficiency}")
    print()

    ws.append([
        expression_form,
        sequential_calculation_time,
        parallel_calculation_time,
        acceleration_factor,
        system_efficiency
    ])


def stats(*expression_forms):
    wb = Workbook()
    ws = wb.active

    ws["A1"] = "Форма виразу"
    ws["B1"] = "Час послідовного обрахунку виразу"
    ws["C1"] = "Час паралельного обрахунку виразу"
    ws["D1"] = "Коефіцієнт прискорення"
    ws["E1"] = "Ефективність"

    for expression_form in expression_forms:
        tree_builder = TreeBuilder(expression_form)
        # tree_builder.print_tree()
        expression_tree_root = tree_builder.building_tree()

        system_characteristics = simulate_parallel_calculation_in_vliw_system(expression_tree_root)
        writing_stats_in_xlsx(ws, expression_form, **system_characteristics)
    else:
        wb.save("stats.xlsx")


if __name__ == "__main__":
    expression = "a*((b+c)/g+(h+i)/g)*d+e*((b+c)/g+(h+i)/g)*d+a*((b+c)/g+(h+i)/g)*f"
    print(f"Expression:\n{expression}")
    print()

    optimized_expression = ExpressionOptimizer(expression).optimizer()

    distributive_transformer = DistributiveTransformator(optimized_expression)
    distributive_forms = distributive_transformer.expression_forms()

    associative_transformator = AssociativeTransformer(optimized_expression)
    associative_forms = associative_transformator.expression_forms()

    stats(expression, *distributive_forms, *associative_forms)
